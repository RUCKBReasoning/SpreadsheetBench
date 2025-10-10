import logging
import os
import shutil
import tempfile
import traceback
import zipfile
from pathlib import Path

import lxml.etree as ET
import openpyxl

# Logger setup
logger = logging.getLogger(__name__)


def ensure_ooxml_compliance(file_path: str) -> str:
    """
    Ensure an Excel file is OOXML compliant by validating and fixing common issues.

    Args:
        file_path: Path to the Excel file to check/fix

    Returns:
        Path to the compliant Excel file with _ooxml suffix (creates if doesn't exist,
        or returns existing compliant file)
    """
    logger.debug(f"Checking OOXML compliance for: {file_path}")

    # Check if file exists
    if not os.path.exists(file_path):
        logger.error(f"File not found: {file_path}")
        return file_path  # Return original path to let openpyxl handle the error

    # Generate compliant file path with _ooxml suffix
    file_dir = os.path.dirname(file_path)
    file_name = os.path.basename(file_path)
    name_without_ext, ext = os.path.splitext(file_name)
    compliant_file_path = os.path.join(file_dir, f"{name_without_ext}_ooxml{ext}")

    # If compliant file already exists, use it
    if os.path.exists(compliant_file_path):
        try:
            test_wb = openpyxl.load_workbook(
                compliant_file_path, read_only=True, data_only=True
            )
            test_wb.close()
            logger.debug(f"Using existing OOXML compliant file: {compliant_file_path}")
            return compliant_file_path
        except Exception as e:
            logger.warning(f"Existing compliant file is invalid, regenerating: {e}")
            # Continue to regenerate the file

    # Check if original file is already compliant
    try:
        test_wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        test_wb.close()
        logger.debug(
            f"File is already OOXML compliant, copying to: {compliant_file_path}"
        )
        shutil.copy2(file_path, compliant_file_path)
        return compliant_file_path
    except Exception as e:
        logger.debug(f"File needs OOXML compliance fixes: {e}")
        # Continue with fixing

    # Create a temporary directory for processing
    temp_dir = tempfile.mkdtemp(prefix="ooxml_fix_")
    try:
        temp_path = Path(temp_dir)

        # Copy the original file to temp location
        temp_excel = temp_path / "workbook.xlsx"
        shutil.copy2(file_path, temp_excel)

        # Extract the Excel file (it's a ZIP archive)
        extract_dir = temp_path / "extracted"
        extract_dir.mkdir()

        try:
            with zipfile.ZipFile(temp_excel, "r") as zip_ref:
                zip_ref.extractall(extract_dir)
        except zipfile.BadZipFile:
            logger.error(f"File is not a valid Excel/ZIP file: {file_path}")
            shutil.rmtree(temp_dir)  # Clean up
            return file_path  # Return original if can't process

        # Fix common OOXML compliance issues
        fixed = False

        # 1. Fix workbook.xml namespace issues and defined names
        workbook_path = extract_dir / "xl" / "workbook.xml"
        if workbook_path.exists():
            fixed |= _fix_workbook_namespaces(workbook_path)
            fixed |= _fix_defined_names(workbook_path)

        # 2. Fix shared strings issues
        shared_strings_path = extract_dir / "xl" / "sharedStrings.xml"
        if shared_strings_path.exists():
            fixed |= _fix_shared_strings(shared_strings_path)

        # 3. Fix worksheet issues
        worksheets_dir = extract_dir / "xl" / "worksheets"
        if worksheets_dir.exists():
            for sheet_file in worksheets_dir.glob("*.xml"):
                fixed |= _fix_worksheet_issues(sheet_file)

        # 4. Remove non-standard extensions
        fixed |= _remove_nonstandard_files(extract_dir)

        if not fixed:
            logger.debug("No fixes applied, file structure seems OK")
            shutil.rmtree(temp_dir)  # Clean up
            return file_path

        # Try to repackage the Excel file directly to the final path
        try:
            logger.info(f"Creating ZIP at: {compliant_file_path}")
            with zipfile.ZipFile(
                compliant_file_path, "w", zipfile.ZIP_DEFLATED
            ) as zip_out:
                for root, dirs, files in os.walk(extract_dir):
                    for file in files:
                        full_file_path = os.path.join(root, file)
                        file_path_in_zip = os.path.relpath(full_file_path, extract_dir)

                        # Normalize path separators for ZIP (always use forward slashes)
                        file_path_in_zip = file_path_in_zip.replace(os.sep, "/")
                        zip_out.write(full_file_path, file_path_in_zip)
        except OSError as e:
            # If direct write fails (e.g., GCP mount doesn't support seek), use temp file
            logger.info(f"Direct ZIP creation failed ({e}), using temp file approach")
            temp_zip = temp_path / "compliant.xlsx"

            with zipfile.ZipFile(temp_zip, "w", zipfile.ZIP_DEFLATED) as zip_out:
                for root, dirs, files in os.walk(extract_dir):
                    for file in files:
                        full_file_path = os.path.join(root, file)
                        file_path_in_zip = os.path.relpath(full_file_path, extract_dir)

                        # Normalize path separators for ZIP (always use forward slashes)
                        file_path_in_zip = file_path_in_zip.replace(os.sep, "/")
                        zip_out.write(full_file_path, file_path_in_zip)

            # Copy from temp to final destination (use copyfileobj for GCP mounts)
            logger.info(f"Copying ZIP to final location: {compliant_file_path}")
            with open(temp_zip, "rb") as src, open(compliant_file_path, "wb") as dst:
                shutil.copyfileobj(src, dst)

        # Clean up temp directory
        shutil.rmtree(temp_dir)

        # Verify the fixed file can be opened
        try:
            verify_wb = openpyxl.load_workbook(
                compliant_file_path, read_only=True, data_only=True
            )
            verify_wb.close()
            logger.debug(
                f"Successfully created OOXML compliant file: {compliant_file_path}"
            )
            return compliant_file_path
        except Exception as verify_error:
            logger.warning(f"Fixed file still has issues: {verify_error}")
            return file_path

    except Exception as e:
        logger.error(f"Error during OOXML compliance fixing: {e}")
        logger.error(f"Full traceback:\n{traceback.format_exc()}")
        # Clean up temp directory on error
        if "temp_dir" in locals() and os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
            except Exception:
                pass
        return file_path  # Return original on error


def _fix_workbook_namespaces(workbook_path: Path) -> bool:
    """Fix namespace issues in workbook.xml"""
    try:
        parser = ET.XMLParser(remove_blank_text=True, recover=True)
        tree = ET.parse(str(workbook_path), parser)
        root = tree.getroot()

        # Register standard OOXML namespaces
        namespaces = {
            "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
            "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
            "mc": "http://schemas.openxmlformats.org/markup-compatibility/2006",
            "x14ac": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac",
        }

        # Remove non-standard namespaces and attributes
        modified = False
        for elem in root.iter():
            # Remove non-standard attributes
            attrs_to_remove = []
            for attr in elem.attrib:
                if attr.startswith("{") and not any(
                    ns in attr for ns in namespaces.values()
                ):
                    attrs_to_remove.append(attr)

            for attr in attrs_to_remove:
                del elem.attrib[attr]
                modified = True

        if modified:
            tree.write(
                str(workbook_path),
                encoding="UTF-8",
                xml_declaration=True,
                pretty_print=True,
            )
            logger.debug("Fixed workbook namespace issues")

        return modified
    except Exception as e:
        logger.warning(f"Could not fix workbook namespaces: {e}")
        logger.warning(f"Traceback:\n{traceback.format_exc()}")
        return False


def _fix_shared_strings(shared_strings_path: Path) -> bool:
    """Fix issues in sharedStrings.xml"""
    try:
        parser = ET.XMLParser(remove_blank_text=True, recover=True)
        tree = ET.parse(str(shared_strings_path), parser)
        root = tree.getroot()

        modified = False

        # Fix encoding issues and invalid characters
        for elem in root.iter():
            if elem.text:
                # Remove invalid XML characters
                cleaned = "".join(
                    char for char in elem.text if ord(char) >= 32 or char in "\t\n\r"
                )
                if cleaned != elem.text:
                    elem.text = cleaned
                    modified = True

        if modified:
            tree.write(
                str(shared_strings_path),
                encoding="UTF-8",
                xml_declaration=True,
                pretty_print=True,
            )
            logger.debug("Fixed shared strings issues")

        return modified
    except Exception as e:
        logger.warning(f"Could not fix shared strings: {e}")
        logger.warning(f"Traceback:\n{traceback.format_exc()}")
        return False


def _fix_worksheet_issues(sheet_path: Path) -> bool:
    """Fix common worksheet compliance issues"""
    try:
        parser = ET.XMLParser(remove_blank_text=True, recover=True)
        tree = ET.parse(str(sheet_path), parser)
        root = tree.getroot()

        modified = False

        # Fix date/time values (ensure they use proper serial format)
        namespace = {
            "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        }

        # Fix date cells
        for cell in root.xpath('.//main:c[@t="d"]', namespaces=namespace):
            value_elem = cell.find("main:v", namespace)
            if value_elem is not None and value_elem.text:
                try:
                    if "T" not in value_elem.text:
                        value_elem.text += "T00:00:00"
                    modified = True
                except Exception:
                    pass

        # Remove invalid formulas
        for formula in root.xpath(
            './/main:f[starts-with(text(), "_xlfn.")]', namespaces=namespace
        ):
            parent = formula.getparent()
            parent.remove(formula)
            modified = True

        if modified:
            tree.write(
                str(sheet_path),
                encoding="UTF-8",
                xml_declaration=True,
                pretty_print=True,
            )
            logger.debug(f"Fixed worksheet issues in {sheet_path.name}")

        return modified
    except Exception as e:
        logger.warning(f"Could not fix worksheet {sheet_path.name}: {e}")
        logger.warning(f"Traceback:\n{traceback.format_exc()}")
        return False


def _fix_defined_names(workbook_path: Path) -> bool:
    """Fix issues with defined names in workbook.xml"""
    try:
        parser = ET.XMLParser(remove_blank_text=True, recover=True)
        tree = ET.parse(str(workbook_path), parser)
        root = tree.getroot()

        modified = False
        namespace = {
            "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        }

        # Find definedNames element
        defined_names = root.find(".//main:definedNames", namespace)
        if defined_names is not None:
            names_to_remove = []

            for defined_name in defined_names.findall("main:definedName", namespace):
                # Remove unsupported attributes
                attrs_to_remove = []
                for attr in defined_name.attrib:
                    # Keep only standard attributes
                    if attr not in [
                        "name",
                        "localSheetId",
                        "hidden",
                        "comment",
                        "customMenu",
                        "description",
                        "help",
                        "statusBar",
                        "vbProcedure",
                        "publishToServer",
                        "workbookParameter",
                        "xlm",
                        "functionGroupId",
                    ]:
                        # Remove non-standard attributes like 'isReadOnly'
                        attrs_to_remove.append(attr)

                for attr in attrs_to_remove:
                    del defined_name.attrib[attr]
                    modified = True
                    logger.debug(
                        f"Removed unsupported attribute '{attr}' from definedName"
                    )

                # Also check if the defined name has invalid references
                if defined_name.text:
                    # Remove defined names with #REF! errors or other issues
                    if "#REF!" in defined_name.text or "#NAME?" in defined_name.text:
                        names_to_remove.append(defined_name)
                        modified = True

            # Remove problematic defined names
            for name in names_to_remove:
                defined_names.remove(name)
                logger.debug(
                    f"Removed problematic definedName: {name.get('name', 'unknown')}"
                )

        if modified:
            tree.write(
                str(workbook_path),
                encoding="UTF-8",
                xml_declaration=True,
                pretty_print=True,
            )
            logger.debug("Fixed defined names issues in workbook.xml")

        return modified
    except Exception as e:
        logger.warning(f"Could not fix defined names: {e}")
        logger.warning(f"Traceback:\n{traceback.format_exc()}")
        return False


def _remove_nonstandard_files(extract_dir: Path) -> bool:
    """Remove non-standard files that may cause compliance issues"""
    modified = False

    # List of non-standard directories/files to remove
    nonstandard_items = [
        "xl/userNames",
        "xl/revisions",
        "xl/richData",
        "xl/metadata.xml",
        "xl/richValueRels",
        "customXml",
    ]

    for item in nonstandard_items:
        item_path = extract_dir / item
        if item_path.exists():
            if item_path.is_dir():
                shutil.rmtree(item_path)
            else:
                item_path.unlink()
            logger.debug(f"Removed non-standard item: {item}")
            modified = True

    # Also update relationships to remove references to deleted items
    rels_path = extract_dir / "xl" / "_rels" / "workbook.xml.rels"
    if rels_path.exists() and modified:
        try:
            parser = ET.XMLParser(remove_blank_text=True, recover=True)
            tree = ET.parse(str(rels_path), parser)
            root = tree.getroot()
            namespace = {
                "rel": "http://schemas.openxmlformats.org/package/2006/relationships"
            }

            # Remove relationships to non-standard parts
            relationships_to_remove = []
            for rel in root.findall("rel:Relationship", namespace):
                target = rel.get("Target")
                if target and any(
                    ns in target
                    for ns in ["userNames", "revisions", "richData", "metadata"]
                ):
                    relationships_to_remove.append(rel)

            for rel in relationships_to_remove:
                root.remove(rel)

            if relationships_to_remove:
                tree.write(
                    str(rels_path),
                    encoding="UTF-8",
                    xml_declaration=True,
                    pretty_print=True,
                )
                logger.debug("Updated relationships after removing non-standard items")
        except Exception as e:
            logger.warning(f"Could not update relationships: {e}")
            logger.warning(f"Traceback:\n{traceback.format_exc()}")

    return modified
