import os
import json
import datetime
import openpyxl
import argparse
import numpy as np
from tqdm import tqdm
from collections import defaultdict
from openpyxl.styles import PatternFill, Font


def datetime_to_float(dt):
    excel_start_date = datetime.datetime(1899, 12, 30)
    delta = dt - excel_start_date
    return delta.days + delta.seconds / 86400.0


def transform_value(v):
    if isinstance(v, (int, float)):
        v = round(float(v), 2)
    elif isinstance(v, datetime.time):
        v = str(v)[:-3]
    elif isinstance(v, datetime.datetime):
        v = round(datetime_to_float(v), 0)
    elif isinstance(v, str):
        try:
            v = round(float(v), 2)
        except ValueError:
            pass
    return v


def compare_cell_value(v1, v2):
    v1 = transform_value(v1)
    v2 = transform_value(v2)
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True
    if type(v1) != type(v2):
        # print(type(v1), type(v2))
        return False
    if v1 == v2:
        return True
    else:
        return False


def _get_color_rgb(color) -> str:
    """Extract RGB value from color object, defaulting to '00000000' if not a string."""
    if color and isinstance(color.rgb, str):
        return color.rgb
    return "00000000"


def _compare_colors(color1, color2) -> bool:
    """Compare two colors using only last 6 characters (RGB), ignoring alpha channel."""
    rgb1 = _get_color_rgb(color1)
    rgb2 = _get_color_rgb(color2)
    return rgb1[-6:] == rgb2[-6:]


def compare_fill_color(fill1, fill2) -> bool:
    """Compare fill colors between two cells."""
    return _compare_colors(fill1.fgColor, fill2.fgColor) and _compare_colors(
        fill1.bgColor, fill2.bgColor
    )


def compare_font_color(font_gt, font_proc) -> bool:
    """Compare font colors between two cells."""
    return _compare_colors(font_gt.color, font_proc.color)


def col_num2name(n):
    """Convert a column number to an Excel column name"""
    name = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        name = chr(65 + remainder) + name
    return name


def col_name2num(name):
    """Convert an Excel column name to a column number"""
    num = 0
    for c in name:
        num = num * 26 + (ord(c) - ord("A") + 1)
    return num


def parse_cell_range(range_str):
    """Parse a range string like 'A1:AB12'"""
    start_cell, end_cell = range_str.split(":")
    start_col, start_row = "", ""
    for char in start_cell:
        if char.isdigit():
            start_row += char
        else:
            start_col += char

    end_col, end_row = "", ""
    for char in end_cell:
        if char.isdigit():
            end_row += char
        else:
            end_col += char

    return (col_name2num(start_col), int(start_row)), (
        col_name2num(end_col),
        int(end_row),
    )


def generate_cell_names(range_str):
    """Generate a list of all cell names in the specified range"""
    if ":" not in range_str:
        return [range_str]
    (start_col, start_row), (end_col, end_row) = parse_cell_range(range_str)
    columns = [col_num2name(i) for i in range(start_col, end_col + 1)]
    cell_names = [
        f"{col}{row}" for col in columns for row in range(start_row, end_row + 1)
    ]
    return cell_names


def cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range):
    if sheet_name not in wb_proc:
        return False, "worksheet not found"
    ws_gt = wb_gt[sheet_name]
    ws_proc = wb_proc[sheet_name]

    cell_names = generate_cell_names(cell_range)

    for cell_name in cell_names:
        cell_gt = ws_gt[cell_name]
        cell_proc = ws_proc[cell_name]

        if not compare_cell_value(cell_gt.value, cell_proc.value):
            msg = f"Value difference at cell {cell_gt.coordinate}: ws_gt has {cell_gt.value},\
                    ws_proc has {cell_proc.value}"
            return False, msg

        # if not compare_fill_color(cell_gt.fill, cell_proc.fill):
        #     msg = f"Fill color difference at cell {cell_gt.coordinate}: ws_gt has {cell_gt.fill.fgColor.rgb},\
        #             ws_proc has {cell_proc.fill.fgColor.rgb}"
        #     return False, msg

        # if not compare_font_color(cell_gt.font, cell_proc.font):
        #     # msg = f"Font color difference at cell {cell_gt.coordinate}: ws_gt has {cell_gt.font.color.rgb},\
        #     #        ws_proc has {cell_proc.font.color.rgb}"
        #     msg = f"Font color difference at cell {cell_gt.coordinate}"
        #     return False, msg

    return True, ""


def compare_workbooks(gt_file, proc_file, instruction_type, answer_position):
    if not os.path.exists(proc_file):
        return False, "File not exist"
    # Open workbooks
    try:
        wb_gt = openpyxl.load_workbook(filename=gt_file, data_only=True)
        wb_proc = openpyxl.load_workbook(filename=proc_file, data_only=True)
    except Exception as e:
        return False, str(e)

    # Initialize report
    result = False
    msg = ""

    sheet_cell_ranges = answer_position.split(",")
    result_list = []
    msg_list = []
    for sheet_cell_range in sheet_cell_ranges:
        if "!" in sheet_cell_range:
            sheet_name, cell_range = sheet_cell_range.split("!")
            sheet_name = sheet_name.lstrip("'").rstrip("'")
        else:
            sheet_name = wb_gt.sheetnames[0]
            cell_range = sheet_cell_range

        # process sheet_name and cell_range
        sheet_name = sheet_name.lstrip("'").rstrip("'")
        cell_range = cell_range.lstrip("'").rstrip("'")

        result, msg = cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range)
        result_list.append(result)
        msg_list.append(msg)

    return all(result_list), ""


def parse_option():
    parser = argparse.ArgumentParser("command line arguments for evaluation.")

    parser.add_argument("--model", type=str, default="llama", help="model name")
    parser.add_argument(
        "--setting",
        type=str,
        default="single",
        help="four setting: single, multi_react_exec, multi_row_exec, multi_row_react_exec",
    )
    parser.add_argument(
        "--dataset", type=str, default="all_data_912", help="dataset name"
    )

    opt = parser.parse_args()

    return opt


def evaluation(opt):
    dataset_path = os.path.abspath(f"../data/{opt.dataset}")
    with open(f"{dataset_path}/dataset.json", "r") as fp:
        dataset = json.load(fp)

    eval_results = []
    for data in tqdm(dataset):
        test_case_results = []
        for test_case_idx in range(3):
            gt_path = f"{dataset_path}/spreadsheet/{data['id']}/{test_case_idx + 1}_{data['id']}_answer.xlsx"
            proc_path = f"{dataset_path}/spreadsheet/{data['id']}/{test_case_idx + 1}_{data['id']}_input.xlsx"
            # proc_path = f"{dataset_path}/outputs/{opt.setting}_{opt.model}/{test_case_idx + 1}_{data['id']}_output.xlsx"
            try:
                result, _ = compare_workbooks(
                    gt_path,
                    proc_path,
                    data["instruction_type"],
                    data["answer_position"],
                )
            except:
                result = False
            test_case_results.append(int(result))
        soft_restriction = test_case_results.count(1) / len(test_case_results)
        hard_restriction = 0 if 0 in test_case_results else 1
        eval_results.append(
            {
                "id": data["id"],
                "instruction_type": data["instruction_type"],
                "test_case_results": test_case_results,
                "soft_restriction": soft_restriction,
                "hard_restriction": hard_restriction,
            }
        )

    with open(f"../outputs/eval_{opt.setting}_{opt.model}.json", "w") as fp:
        json.dump(eval_results, fp, indent=4)


if __name__ == "__main__":
    opt = parse_option()
    print(opt)

    evaluation(opt)
