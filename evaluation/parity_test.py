"""
Parity test: verify that LibreOffice recalculation produces the same
evaluation results as the original win32com+Excel pipeline.

Strategy:
1. Run evaluation on raw (unrecalculated) answer files — record "before" results
2. Copy answer files to a temp location, run LibreOffice recalculation on copies
3. Run evaluation on recalculated answer files — record "after" results
4. Compare: identify tasks where recalculation changed the outcome
5. Analyze: check for formula cells with None values before recalculation
"""

import os
import sys
import json
import shutil
import subprocess
import datetime
import openpyxl
from tqdm import tqdm
from collections import defaultdict

DATASET_DIR = os.path.abspath("../data/sample_data_200")
SPREADSHEET_DIR = os.path.join(DATASET_DIR, "spreadsheet")
RECALC_DIR = os.path.abspath("../data/sample_data_200_recalculated")
SOFFICE = "/Applications/LibreOffice.app/Contents/MacOS/soffice"


# --- Evaluation logic (copied from evaluation.py to keep self-contained) ---

def datetime_to_float(dt):
    excel_start_date = datetime.datetime(1899, 12, 30)
    delta = dt - excel_start_date
    return delta.days + delta.seconds / 86400.0


def transform_value(v):
    if isinstance(v, (int, float)):
        v = round(float(v), 2)
    elif isinstance(v, datetime.time):
        v = str(v)[:-3]
    elif isinstance(v, datetime.datetime):
        v = round(datetime_to_float(v), 0)
    elif isinstance(v, str):
        try:
            v = round(float(v), 2)
        except ValueError:
            pass
    return v


def compare_cell_value(v1, v2):
    v1 = transform_value(v1)
    v2 = transform_value(v2)
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True
    if type(v1) != type(v2):
        return False
    return v1 == v2


def col_num2name(n):
    name = ''
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        name = chr(65 + remainder) + name
    return name


def col_name2num(name):
    num = 0
    for c in name:
        num = num * 26 + (ord(c) - ord('A') + 1)
    return num


def parse_cell_range(range_str):
    start_cell, end_cell = range_str.split(':')
    start_col, start_row = '', ''
    for char in start_cell:
        if char.isdigit():
            start_row += char
        else:
            start_col += char
    end_col, end_row = '', ''
    for char in end_cell:
        if char.isdigit():
            end_row += char
        else:
            end_col += char
    return (col_name2num(start_col), int(start_row)), (col_name2num(end_col), int(end_row))


def generate_cell_names(range_str):
    if ':' not in range_str:
        return [range_str]
    (start_col, start_row), (end_col, end_row) = parse_cell_range(range_str)
    columns = [col_num2name(i) for i in range(start_col, end_col + 1)]
    return [f"{col}{row}" for col in columns for row in range(start_row, end_row + 1)]


def cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range):
    if sheet_name not in wb_proc:
        return False, "worksheet not found"
    ws_gt = wb_gt[sheet_name]
    ws_proc = wb_proc[sheet_name]
    for cell_name in generate_cell_names(cell_range):
        cell_gt = ws_gt[cell_name]
        cell_proc = ws_proc[cell_name]
        if not compare_cell_value(cell_gt.value, cell_proc.value):
            return False, f"Value diff at {cell_name}: gt={cell_gt.value!r}, proc={cell_proc.value!r}"
    return True, ""


def compare_workbooks(gt_file, proc_file, answer_position):
    if not os.path.exists(proc_file):
        return False, "File not exist"
    try:
        wb_gt = openpyxl.load_workbook(filename=gt_file, data_only=True)
        wb_proc = openpyxl.load_workbook(filename=proc_file, data_only=True)
    except Exception as e:
        return False, str(e)

    sheet_cell_ranges = answer_position.split(',')
    for scr in sheet_cell_ranges:
        if '!' in scr:
            sheet_name, cell_range = scr.split('!')
            sheet_name = sheet_name.strip("'")
        else:
            sheet_name = wb_gt.sheetnames[0]
            cell_range = scr
        cell_range = cell_range.strip("'")
        result, msg = cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range)
        if not result:
            return False, msg
    return True, ""


# --- Formula analysis ---

def check_formulas_in_file(filepath):
    """Check if a file has formula cells and whether they have cached values."""
    try:
        # Load with formulas visible
        wb_formula = openpyxl.load_workbook(filepath, data_only=False)
        # Load with data_only to see cached values
        wb_data = openpyxl.load_workbook(filepath, data_only=True)
    except Exception:
        return 0, 0, 0

    total_formulas = 0
    cached = 0
    uncached = 0

    for sheet_name in wb_formula.sheetnames:
        ws_f = wb_formula[sheet_name]
        ws_d = wb_data[sheet_name]
        for row in ws_f.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    total_formulas += 1
                    data_cell = ws_d[cell.coordinate]
                    if data_cell.value is not None:
                        cached += 1
                    else:
                        uncached += 1

    return total_formulas, cached, uncached


# --- LibreOffice recalculation ---

def recalculate_file(filepath):
    """Recalculate a single xlsx file using LibreOffice."""
    import tempfile
    dirname = os.path.dirname(filepath)
    basename = os.path.basename(filepath)
    name, ext = os.path.splitext(basename)

    with tempfile.TemporaryDirectory() as tmpdir:
        result = subprocess.run(
            [SOFFICE, "--headless", "--calc",
             "--convert-to", "xlsx:Calc MS Excel 2007 XML",
             "--outdir", tmpdir, filepath],
            capture_output=True, text=True, timeout=120,
        )
        if result.returncode != 0:
            return False
        converted = os.path.join(tmpdir, name + ".xlsx")
        if not os.path.isfile(converted):
            return False
        shutil.move(converted, os.path.join(dirname, name + ".xlsx"))
        if ext.lower() == ".xls":
            os.remove(filepath)
        return True


def run_evaluation(dataset, spreadsheet_dir):
    """Run evaluation and return per-task results."""
    results = {}
    for data in dataset:
        task_id = data['id']
        test_case_results = []
        for tc in range(3):
            gt_path = os.path.join(spreadsheet_dir, str(task_id),
                                   f"{tc + 1}_{task_id}_answer.xlsx")
            proc_path = os.path.join(spreadsheet_dir, str(task_id),
                                     f"{tc + 1}_{task_id}_input.xlsx")
            try:
                result, msg = compare_workbooks(gt_path, proc_path, data['answer_position'])
            except Exception:
                result = False
            test_case_results.append(int(result))

        results[task_id] = {
            'test_case_results': test_case_results,
            'soft': test_case_results.count(1) / len(test_case_results),
            'hard': 0 if 0 in test_case_results else 1,
        }
    return results


def main():
    with open(os.path.join(DATASET_DIR, "dataset.json")) as f:
        dataset = json.load(f)
    print(f"Loaded {len(dataset)} tasks from dataset.json\n")

    # ---- Phase 1: Analyze formula status in answer files ----
    print("=" * 60)
    print("PHASE 1: Analyzing formula cells in answer files")
    print("=" * 60)

    files_with_formulas = 0
    files_with_uncached = 0
    total_formula_cells = 0
    total_uncached_cells = 0

    for data in tqdm(dataset, desc="Scanning formulas"):
        task_id = data['id']
        for tc in range(3):
            answer_path = os.path.join(SPREADSHEET_DIR, str(task_id),
                                        f"{tc + 1}_{task_id}_answer.xlsx")
            if os.path.exists(answer_path):
                formulas, cached, uncached = check_formulas_in_file(answer_path)
                if formulas > 0:
                    files_with_formulas += 1
                    total_formula_cells += formulas
                    if uncached > 0:
                        files_with_uncached += 1
                        total_uncached_cells += uncached

    print(f"\nFormula analysis:")
    print(f"  Files with formulas:         {files_with_formulas}")
    print(f"  Files with UNCACHED formulas: {files_with_uncached}")
    print(f"  Total formula cells:         {total_formula_cells}")
    print(f"  Uncached formula cells:      {total_uncached_cells}")

    # ---- Phase 2: Run evaluation BEFORE recalculation ----
    print(f"\n{'=' * 60}")
    print("PHASE 2: Evaluation BEFORE recalculation (original files)")
    print("=" * 60)

    results_before = run_evaluation(dataset, SPREADSHEET_DIR)
    before_hard = sum(r['hard'] for r in results_before.values())
    before_soft = sum(r['soft'] for r in results_before.values())
    print(f"  Hard accuracy: {before_hard}/{len(results_before)} ({100*before_hard/len(results_before):.1f}%)")
    print(f"  Soft accuracy: {before_soft/len(results_before):.4f}")

    # ---- Phase 3: Copy and recalculate with LibreOffice ----
    print(f"\n{'=' * 60}")
    print("PHASE 3: Recalculating with LibreOffice")
    print("=" * 60)

    # Copy spreadsheet dir
    recalc_spreadsheet_dir = os.path.join(RECALC_DIR, "spreadsheet")
    if os.path.exists(RECALC_DIR):
        shutil.rmtree(RECALC_DIR)
    os.makedirs(RECALC_DIR)
    shutil.copytree(SPREADSHEET_DIR, recalc_spreadsheet_dir)

    # Recalculate all answer AND input xlsx files
    all_xlsx = []
    for root, dirs, files in os.walk(recalc_spreadsheet_dir):
        for f in files:
            if f.endswith(('.xlsx', '.xls')):
                all_xlsx.append(os.path.join(root, f))

    print(f"  Found {len(all_xlsx)} spreadsheet files to recalculate")
    success = 0
    failed = 0
    for filepath in tqdm(all_xlsx, desc="Recalculating"):
        if recalculate_file(filepath):
            success += 1
        else:
            failed += 1
            print(f"  FAILED: {filepath}")

    print(f"  Recalculated: {success} succeeded, {failed} failed")

    # ---- Phase 4: Run evaluation AFTER recalculation ----
    print(f"\n{'=' * 60}")
    print("PHASE 4: Evaluation AFTER recalculation (LibreOffice)")
    print("=" * 60)

    results_after = run_evaluation(dataset, recalc_spreadsheet_dir)
    after_hard = sum(r['hard'] for r in results_after.values())
    after_soft = sum(r['soft'] for r in results_after.values())
    print(f"  Hard accuracy: {after_hard}/{len(results_after)} ({100*after_hard/len(results_after):.1f}%)")
    print(f"  Soft accuracy: {after_soft/len(results_after):.4f}")

    # ---- Phase 5: Compare before vs after ----
    print(f"\n{'=' * 60}")
    print("PHASE 5: Parity comparison")
    print("=" * 60)

    improved = []
    regressed = []
    unchanged = 0

    for task_id in results_before:
        before = results_before[task_id]
        after = results_after[task_id]
        if before['hard'] < after['hard']:
            improved.append(task_id)
        elif before['hard'] > after['hard']:
            regressed.append(task_id)
        else:
            # Check soft score changes
            if before['soft'] < after['soft']:
                improved.append(task_id)
            elif before['soft'] > after['soft']:
                regressed.append(task_id)
            else:
                unchanged += 1

    print(f"\n  Tasks improved by recalculation:  {len(improved)}")
    print(f"  Tasks regressed by recalculation: {len(regressed)}")
    print(f"  Tasks unchanged:                  {unchanged}")

    if regressed:
        print(f"\n  REGRESSIONS (need investigation):")
        for tid in regressed[:20]:
            b = results_before[tid]
            a = results_after[tid]
            print(f"    Task {tid}: before={b['test_case_results']} after={a['test_case_results']}")

    if improved:
        print(f"\n  IMPROVEMENTS (formulas now correctly evaluated):")
        for tid in improved[:20]:
            b = results_before[tid]
            a = results_after[tid]
            print(f"    Task {tid}: before={b['test_case_results']} after={a['test_case_results']}")

    # ---- Summary ----
    print(f"\n{'=' * 60}")
    print("SUMMARY")
    print("=" * 60)
    print(f"  Before recalc — Hard: {before_hard}/{len(results_before)} | Soft: {before_soft/len(results_before):.4f}")
    print(f"  After recalc  — Hard: {after_hard}/{len(results_after)}  | Soft: {after_soft/len(results_after):.4f}")
    print(f"  Regressions: {len(regressed)}")
    if len(regressed) == 0:
        print(f"\n  ** NO REGRESSIONS — LibreOffice is safe to use as a drop-in replacement **")
    else:
        print(f"\n  ** REGRESSIONS DETECTED — investigation needed **")


if __name__ == "__main__":
    main()
